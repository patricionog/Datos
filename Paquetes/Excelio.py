from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from Listio import *
import pandas as pd

def Get_Sheets_And_Sheetnames(Book):
    Sheet_Names = Book.sheetnames 
    Sheets = []
    for Sheet in Sheets:
        Current_Sheet = Book[Sheet_Names[Sheet]]
        Sheets.append(Current_Sheet)
    return Sheet_Names, Sheets

def Get_Range_Values(Cell1, Cell2, Sheet):
    Range = Sheet[Cell1:Cell2]
    Range = Convert_List_Of_Tuples_To_List_Of_List(Range)
    Range_Values = []

    for Row in Range:
        Row_Values = []
        for Cell in Row:
            Row_Values.append(Cell.value)
        Range_Values.append(Row_Values)
    return Range_Values

def Verify_Cell_Value(Value, Sheet, Column, Row):
    return Value == Sheet[Column][Row]

def Get_Cell_Value(Value, Sheet, Cell = None, Column = None, Row = None):
    if Cell is None and Column is None and Row is None:
        raise KeyError(
            f'Tenés que insertar una celda. Ejemplo: "A1". \n', 
            'O podés ingresar la columna y la fila, de esta manera: \n',
            'Column = 1, Row = 2')
    
    if Cell:
        return Value == Sheet[Cell].value
    
    else:
        return Sheet.cell(row=Row, column=Column).value

def Get_Length_And_Width(Sheet):
    Length = Sheet.max_row  
    Width = Sheet.max_column
    return Length, Width

def Merge_Sheets(Book, Name):
    New_Doc = Workbook()
    New_Sheet = New_Doc.create_sheet("Result")
    Sheet_Names, Sheets = Get_Sheets_And_Sheetnames(Book)
    Total_Length = 1

    for Sheet in Sheets:
        Sheet_Length, Sheet_Width = Get_Length_And_Width(Sheets[Sheet])    
        for Row in range(Total_Length, Total_Length + Sheet_Length):
            for Column in range(1, Sheet_Width):
                Cell = Sheets[Sheet][chr(Column + 64) + str(Row - Total_Length + 1)].value
                New_Sheet.cell(row=Row, column=Column, value=Cell)
        Total_Length += Sheet_Length

def Formating_Book(Path, Sheet_Name=None):    
    Book = load_workbook(Path)
    
    if Sheet_Name == None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]

    # Filter.
    Sheet.auto_filter.ref = Sheet.dimensions

    Align_Format = Alignment(horizontal='center', vertical='center')

    Border_Format = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'), bottom=Side(style='thin'))
    
    for Column in range(1, Sheet.max_column + 1):
        for Row in range(1, Sheet.max_row + 1):
            Cell = Sheet.cell(row=Row, column=Column)
            Cell.alignment = Align_Format
            Cell.border = Border_Format

    Book.save(Path)

def Convert_df_To_Excel(df, Path, Filename):    
    Path = Path + Filename
    Formating_Book(Path)

def Immobilize_Panes(Path, Cell, Sheet_Name=None):
    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        if Sheet_Name not in Book.sheetnames:
            raise ValueError(f"Sheet '{Sheet_Name}' does not exist in the workbook.")
        Sheet = Book[Sheet_Name]

    Sheet.freeze_panes = Cell
    Book.save(Path)

def Replace_Character_In_Range_Of_Excel_Book(Path, Old_Character, New_Character, Min_Col, Min_Row, Max_Col=None, Max_Row=None, Sheet_Name=None):
    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
    
    if Max_Col is None:
        Max_Col = Sheet.max_column
    if Max_Row is None:
        Max_Row = Sheet.max_row
    
    for Row in Sheet.iter_rows(min_row=Min_Row, min_col=Min_Col, max_row=Max_Row, max_col=Max_Col):  
        for Cell in Row:
            Content = Cell.value
            if Content:
                Cell.value = Content.replace(Old_Character, New_Character)
    
    Book.save(Path)

def Delete_Columns_In_Excel_Book(Path, Columns, Sheet_Name=None):
    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
    
    Row1 = Sheet[1]
    
    Mapping = {Cell.value: Cell.column for Cell in Row1}
    
    Column_Index = [Mapping[i] for i in Columns if i in Mapping]
     
    for i in sorted(Column_Index, reverse=True):
        Sheet.delete_cols(i)
    
    Book.save(Path)

def Adjust_Column_Width(Path, Min_Column, Max_Column, Width, Sheet_Name=None):
    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
        
    for Column in range(Min_Column, Max_Column + 1):
        Sheet.column_dimensions[chr(64 + Column)].width = Width
    
    Book.save(Path)

def Create_Excel_With_Multiple_Data_Frames(Path, df_List, Sheet_Names_List=None):
    if Sheet_Names_List is None:
        Sheet_Names_List = [f'Sheet{i+1}' for i in range(len(df_List))]

    elif len(df_List) != len(Sheet_Names_List):
        raise ValueError("The list of sheet names must be the same length as the list of DataFrames.")

    with pd.ExcelWriter(Path) as Writer:
        for df, Sheet_Name in zip(df_List, Sheet_Names_List):
            df.to_excel(Writer, sheet_name=Sheet_Name, index=False)